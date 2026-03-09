#!/usr/bin/env python3
"""
migrate_xlsx.py — Freshorino POS
Migrate your store.xlsx data into the new SQLite database.

Usage:
    python migrate_xlsx.py store.xlsx [store.db]
"""
import sqlite3, sys, os

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

sys.path.insert(0, os.path.dirname(__file__) or ".")
XLSX = sys.argv[1] if len(sys.argv) > 1 else "store.xlsx"
os.environ["DB_PATH"] = sys.argv[2] if len(sys.argv) > 2 else "store.db"

from database import init_db, get_connection
init_db()

if not os.path.exists(XLSX):
    print(f"ERROR: {XLSX} not found"); sys.exit(1)

wb = openpyxl.load_workbook(XLSX)
conn = get_connection()
c = conn.cursor()

def col(row, hdr, name):
    try: return row[hdr.index(name)]
    except: return None

total_imported = 0

for table, col_map in [
    ("products", {"id":"id","name":"name","barcode":"barcode","price":"price",
                  "purchase_price":"purchase_price","quantity":"quantity",
                  "unit_type":"unit_type","expiry_date":"expiry_date"}),
]:
    if table not in wb.sheetnames: continue
    ws = wb[table]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    imp = skp = 0
    for row in rows[1:]:
        pid = col(row, hdr, "id")
        name = (col(row, hdr, "name") or "").strip()
        if not pid or not name: continue
        if c.execute("SELECT 1 FROM products WHERE id=?", (pid,)).fetchone():
            skp += 1; continue
        c.execute("""
            INSERT OR IGNORE INTO products
            (id,name,name_norm,barcode,price,purchase_price,quantity,unit_type,expiry_date,is_quick)
            VALUES(?,?,?,?,?,?,?,?,?,0)
        """, (pid, name, name.casefold(),
              col(row, hdr, "barcode") or None,
              float(col(row, hdr, "price") or 0),
              float(col(row, hdr, "purchase_price") or 0),
              float(col(row, hdr, "quantity") or 0),
              col(row, hdr, "unit_type") or "piece",
              col(row, hdr, "expiry_date") or None))
        imp += 1
    conn.commit()
    max_id = c.execute(f"SELECT MAX(id) FROM {table}").fetchone()[0] or 0
    c.execute(f"INSERT OR REPLACE INTO sqlite_sequence(name,seq) VALUES(?,?)", (table, max_id))
    conn.commit()
    print(f"  ✅ {table}: {imp} imported, {skp} skipped")
    total_imported += imp

# SALES
if "sales" in wb.sheetnames:
    ws = wb["sales"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    imp = skp = 0
    for row in rows[1:]:
        sid = col(row, hdr, "id")
        if not sid: continue
        if c.execute("SELECT 1 FROM sales WHERE id=?", (sid,)).fetchone(): skp+=1; continue
        c.execute("INSERT OR IGNORE INTO sales(id,date,total,discount_percent,discount_amount) VALUES(?,?,?,?,?)",
            (sid, str(col(row,hdr,"date") or "")[:19],
             float(col(row,hdr,"total") or 0),
             float(col(row,hdr,"discount_percent") or 0),
             float(col(row,hdr,"discount_amount") or 0)))
        imp += 1
    conn.commit()
    max_id = c.execute("SELECT MAX(id) FROM sales").fetchone()[0] or 0
    c.execute("INSERT OR REPLACE INTO sqlite_sequence(name,seq) VALUES('sales',?)", (max_id,))
    conn.commit()
    print(f"  ✅ sales: {imp} imported, {skp} skipped")
    total_imported += imp

# SALE_ITEMS
if "sale_items" in wb.sheetnames:
    ws = wb["sale_items"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    imp = skp = 0
    for row in rows[1:]:
        iid = col(row, hdr, "id")
        if not iid: continue
        if c.execute("SELECT 1 FROM sale_items WHERE id=?", (iid,)).fetchone(): skp+=1; continue
        sale_id = col(row, hdr, "sale_id")
        product_id = col(row, hdr, "product_id")
        if not c.execute("SELECT 1 FROM sales WHERE id=?", (sale_id,)).fetchone(): skp+=1; continue
        if not c.execute("SELECT 1 FROM products WHERE id=?", (product_id,)).fetchone(): skp+=1; continue
        pname = c.execute("SELECT name FROM products WHERE id=?", (product_id,)).fetchone()
        c.execute("INSERT OR IGNORE INTO sale_items(id,sale_id,product_id,product_name,quantity,unit_type,unit_price,cost_price,line_total) VALUES(?,?,?,?,?,?,?,?,?)",
            (iid, sale_id, product_id, pname[0] if pname else "",
             float(col(row,hdr,"quantity") or 0), col(row,hdr,"unit_type") or "piece",
             float(col(row,hdr,"unit_price") or 0), float(col(row,hdr,"cost_price") or 0),
             float(col(row,hdr,"line_total") or 0)))
        imp += 1
    conn.commit()
    max_id = c.execute("SELECT MAX(id) FROM sale_items").fetchone()[0] or 0
    c.execute("INSERT OR REPLACE INTO sqlite_sequence(name,seq) VALUES('sale_items',?)", (max_id,))
    conn.commit()
    print(f"  ✅ sale_items: {imp} imported, {skp} skipped")
    total_imported += imp

# WASTE
if "waste" in wb.sheetnames:
    ws = wb["waste"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    imp = skp = 0
    for row in rows[1:]:
        wid = col(row, hdr, "id")
        if not wid: continue
        if c.execute("SELECT 1 FROM waste WHERE id=?", (wid,)).fetchone(): skp+=1; continue
        product_id = col(row, hdr, "product_id")
        if not c.execute("SELECT 1 FROM products WHERE id=?", (product_id,)).fetchone(): skp+=1; continue
        c.execute("INSERT OR IGNORE INTO waste(id,date,product_id,quantity,unit_type,reason) VALUES(?,?,?,?,?,?)",
            (wid, str(col(row,hdr,"date") or "")[:19], product_id,
             float(col(row,hdr,"quantity") or 0),
             col(row,hdr,"unit_type") or "piece",
             col(row,hdr,"reason") or None))
        imp += 1
    conn.commit()
    max_id = c.execute("SELECT MAX(id) FROM waste").fetchone()[0] or 0
    c.execute("INSERT OR REPLACE INTO sqlite_sequence(name,seq) VALUES('waste',?)", (max_id,))
    conn.commit()
    print(f"  ✅ waste: {imp} imported, {skp} skipped")
    total_imported += imp

conn.close()
print(f"\n🎉 Done! {total_imported} records migrated.")
print("   Default login: admin / admin123")
print("   ⚠️  CHANGE PASSWORD after first login at /admin/users")
